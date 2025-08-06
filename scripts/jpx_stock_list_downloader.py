#!/usr/bin/env python3
"""
JPX銘柄リストダウンローダー

日本取引所グループ（JPX）から最新の上場銘柄一覧をダウンロードする
スタンドアロンスクリプト。

機能:
- JPXウェブサイトから銘柄一覧の自動ダウンロード
- 複数フォーマット対応（CSV、Excel）
- ファイル形式の自動変換
- データ品質検証
- 継続的データ更新対応

Usage:
    python scripts/jpx_stock_list_downloader.py
    python scripts/jpx_stock_list_downloader.py --output data/stocks.csv
    python scripts/jpx_stock_list_downloader.py --format csv --validate
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any
from urllib.parse import urlparse, urljoin

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# オプション依存関係
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# プロジェクトルートをパスに追加
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class JPXStockListDownloader:
    """JPX銘柄リストダウンローダー"""

    # JPXの公開データURL（2024年現在の推定）
    JPX_BASE_URL = "https://www.jpx.co.jp"

    # 可能性のある銘柄リストURL
    POSSIBLE_URLS = [
        # 最新の銘柄一覧（CSV形式）
        "https://www.jpx.co.jp/markets/statistics-equities/misc/tvdivq0000001vg2-att/data_j.csv",

        # Excel形式
        "https://www.jpx.co.jp/markets/statistics-equities/misc/tvdivq0000001vg2-att/data_j.xls",
        "https://www.jpx.co.jp/markets/statistics-equities/misc/tvdivq0000001vg2-att/data_j.xlsx",

        # 市場別データ
        "https://www.jpx.co.jp/markets/statistics-equities/misc/tvdivq0000001vg2-att/listed_stocks_prime.csv",
        "https://www.jpx.co.jp/markets/statistics-equities/misc/tvdivq0000001vg2-att/listed_stocks_standard.csv",
        "https://www.jpx.co.jp/markets/statistics-equities/misc/tvdivq0000001vg2-att/listed_stocks_growth.csv",

        # 代替URL（定期的に更新される可能性があります）
        "https://www.jpx.co.jp/listing/stocks/listing/", # ページからリンクを抽出
    ]

    def __init__(self, timeout: int = 30, retries: int = 3, user_agent: Optional[str] = None):
        """
        Args:
            timeout: リクエストタイムアウト（秒）
            retries: 再試行回数
            user_agent: カスタムUser-Agent
        """
        self.timeout = timeout
        self.session = self._create_session(retries)

        if user_agent:
            self.session.headers['User-Agent'] = user_agent

        self.downloaded_files: List[Path] = []
        self.failed_urls: List[str] = []

    def _create_session(self, retries: int) -> requests.Session:
        """HTTPセッションを作成"""
        session = requests.Session()

        # リトライ戦略
        retry_strategy = Retry(
            total=retries,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET"]
        )

        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("http://", adapter)
        session.mount("https://", adapter)

        # デフォルトヘッダー
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'ja,en-US;q=0.7,en;q=0.3',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        })

        return session

    def download_all_available_formats(self, output_dir: Path) -> List[Path]:
        """
        利用可能なすべての形式をダウンロード

        Args:
            output_dir: 出力ディレクトリ

        Returns:
            ダウンロードされたファイルのリスト
        """
        output_dir.mkdir(parents=True, exist_ok=True)

        logger.info(f"JPX銘柄リストのダウンロード開始: {len(self.POSSIBLE_URLS)} URLを試行")

        for url in self.POSSIBLE_URLS:
            try:
                if self._is_page_url(url):
                    # HTMLページから実際のダウンロードリンクを抽出
                    download_urls = self._extract_download_links(url)
                    for download_url in download_urls:
                        self._download_single_file(download_url, output_dir)
                else:
                    # 直接ダウンロード
                    self._download_single_file(url, output_dir)

            except Exception as e:
                logger.warning(f"URL処理エラー {url}: {e}")
                self.failed_urls.append(url)
                continue

        logger.info(f"ダウンロード完了: {len(self.downloaded_files)}ファイル成功、{len(self.failed_urls)}URL失敗")

        return self.downloaded_files

    def _is_page_url(self, url: str) -> bool:
        """URLがHTMLページかどうかを判定"""
        return not any(url.endswith(ext) for ext in ['.csv', '.xls', '.xlsx', '.json'])

    def _extract_download_links(self, page_url: str) -> List[str]:
        """HTMLページからダウンロードリンクを抽出"""
        try:
            response = self.session.get(page_url, timeout=self.timeout)
            response.raise_for_status()

            html_content = response.text
            download_links = []

            # 簡単なパターンマッチング（より洗練された解析にはBeautifulSoupを使用）
            import re

            # CSV、Excelファイルへのリンクを検索
            patterns = [
                r'href="([^"]*\.csv[^"]*)"',
                r'href="([^"]*\.xls[x]?[^"]*)"',
                r'href="([^"]*data[^"]*\.csv[^"]*)"',
                r'href="([^"]*銘柄[^"]*\.(csv|xlsx?)[^"]*)"',
            ]

            for pattern in patterns:
                matches = re.finditer(pattern, html_content, re.IGNORECASE)
                for match in matches:
                    link = match.group(1)
                    # 相対URLを絶対URLに変換
                    absolute_link = urljoin(page_url, link)
                    if absolute_link not in download_links:
                        download_links.append(absolute_link)

            logger.info(f"ページ {page_url} から {len(download_links)} 個のダウンロードリンクを発見")
            return download_links

        except Exception as e:
            logger.error(f"ダウンロードリンク抽出エラー {page_url}: {e}")
            return []

    def _download_single_file(self, url: str, output_dir: Path) -> Optional[Path]:
        """単一ファイルをダウンロード"""
        try:
            logger.info(f"ダウンロード試行: {url}")

            response = self.session.get(url, timeout=self.timeout, stream=True)
            response.raise_for_status()

            # ファイル名を決定
            filename = self._determine_filename(url, response)
            output_path = output_dir / filename

            # ファイルサイズチェック
            content_length = response.headers.get('content-length')
            if content_length:
                size_mb = int(content_length) / (1024 * 1024)
                logger.info(f"ファイルサイズ: {size_mb:.2f} MB")

                # 異常に大きいか小さいファイルをチェック
                if size_mb > 100:  # 100MB以上
                    logger.warning(f"ファイルサイズが大きすぎます: {size_mb:.2f} MB")
                elif size_mb < 0.001:  # 1KB未満
                    logger.warning(f"ファイルサイズが小さすぎます: {size_mb:.2f} MB")

            # ファイル保存
            with open(output_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)

            # ダウンロード成功検証
            if self._validate_downloaded_file(output_path):
                self.downloaded_files.append(output_path)
                logger.info(f"ダウンロード成功: {output_path}")
                return output_path
            else:
                logger.error(f"ダウンロードファイルの検証に失敗: {output_path}")
                output_path.unlink()  # 無効なファイルを削除
                return None

        except Exception as e:
            logger.error(f"ダウンロードエラー {url}: {e}")
            self.failed_urls.append(url)
            return None

    def _determine_filename(self, url: str, response: requests.Response) -> str:
        """ファイル名を決定"""
        # Content-Dispositionヘッダーからファイル名を抽出
        content_disposition = response.headers.get('content-disposition', '')
        if 'filename=' in content_disposition:
            import re
            filename_match = re.search(r'filename[*]?=([^;]+)', content_disposition)
            if filename_match:
                filename = filename_match.group(1).strip('"\'')
                return filename

        # URLからファイル名を抽出
        url_path = urlparse(url).path
        if url_path and '.' in url_path.split('/')[-1]:
            filename = url_path.split('/')[-1]
            return filename

        # デフォルトファイル名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Content-Typeからファイル拡張子を推定
        content_type = response.headers.get('content-type', '').lower()
        if 'csv' in content_type:
            ext = 'csv'
        elif 'excel' in content_type or 'spreadsheet' in content_type:
            ext = 'xlsx'
        elif 'json' in content_type:
            ext = 'json'
        else:
            ext = 'dat'

        return f"jpx_stocks_{timestamp}.{ext}"

    def _validate_downloaded_file(self, file_path: Path) -> bool:
        """ダウンロードファイルを検証"""
        try:
            # ファイルサイズチェック
            if file_path.stat().st_size == 0:
                logger.error(f"ファイルサイズが0バイト: {file_path}")
                return False

            # ファイル形式別の基本検証
            if file_path.suffix.lower() == '.csv':
                return self._validate_csv_file(file_path)
            elif file_path.suffix.lower() in ['.xls', '.xlsx']:
                return self._validate_excel_file(file_path)
            else:
                # その他のファイル形式は基本的なサイズチェックのみ
                return file_path.stat().st_size > 100  # 100バイト以上

        except Exception as e:
            logger.error(f"ファイル検証エラー {file_path}: {e}")
            return False

    def _validate_csv_file(self, file_path: Path) -> bool:
        """CSVファイルを検証"""
        try:
            # エンコーディング検出
            encodings = ['utf-8', 'shift_jis', 'cp932', 'utf-8-sig']

            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        # 最初の数行を読んで構造を確認
                        lines = []
                        for i, line in enumerate(f):
                            lines.append(line.strip())
                            if i >= 5:  # 最初の6行をチェック
                                break

                        if len(lines) < 2:
                            logger.warning(f"CSVファイルの行数が少なすぎます: {len(lines)}行")
                            return False

                        # ヘッダー行の存在確認
                        header = lines[0]
                        if ',' not in header and '\t' not in header:
                            logger.warning(f"CSVファイルに適切な区切り文字が見つかりません")
                            return False

                        # データ行の確認
                        data_lines = [line for line in lines[1:] if line]
                        if len(data_lines) < 1:
                            logger.warning(f"CSVファイルにデータ行がありません")
                            return False

                        logger.info(f"CSV検証成功: {len(data_lines)}行のデータ、エンコーディング={encoding}")
                        return True

                except UnicodeDecodeError:
                    continue

            logger.error(f"CSVファイルを適切なエンコーディングで読み込めませんでした")
            return False

        except Exception as e:
            logger.error(f"CSV検証エラー: {e}")
            return False

    def _validate_excel_file(self, file_path: Path) -> bool:
        """Excelファイルを検証"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxlが利用できないため、Excelファイルの詳細検証をスキップします")
            return True

        try:
            import openpyxl

            workbook = openpyxl.load_workbook(file_path, read_only=True)

            if not workbook.sheetnames:
                logger.error("Excelファイルにシートがありません")
                return False

            # 最初のシートを確認
            sheet = workbook[workbook.sheetnames[0]]

            if sheet.max_row < 2:
                logger.warning(f"Excelシートの行数が少なすぎます: {sheet.max_row}行")
                return False

            if sheet.max_column < 2:
                logger.warning(f"Excelシートの列数が少なすぎます: {sheet.max_column}列")
                return False

            logger.info(f"Excel検証成功: {sheet.max_row}行 x {sheet.max_column}列")
            workbook.close()
            return True

        except Exception as e:
            logger.error(f"Excel検証エラー: {e}")
            return False

    def convert_to_csv(self, input_file: Path, output_file: Optional[Path] = None) -> Optional[Path]:
        """ExcelファイルをCSVに変換"""
        if not PANDAS_AVAILABLE:
            logger.error("pandasが利用できないため、Excel->CSV変換ができません")
            return None

        try:
            if input_file.suffix.lower() not in ['.xls', '.xlsx']:
                logger.error(f"サポートされていないファイル形式: {input_file}")
                return None

            if output_file is None:
                output_file = input_file.with_suffix('.csv')

            # Excelファイルを読み込み
            df = pd.read_excel(input_file, sheet_name=0)

            # CSVとして保存
            df.to_csv(output_file, index=False, encoding='utf-8-sig')

            logger.info(f"Excel->CSV変換成功: {input_file} -> {output_file}")
            return output_file

        except Exception as e:
            logger.error(f"Excel->CSV変換エラー: {e}")
            return None

    def get_download_summary(self) -> Dict[str, Any]:
        """ダウンロード結果のサマリーを取得"""
        return {
            'downloaded_files': [str(f) for f in self.downloaded_files],
            'failed_urls': self.failed_urls,
            'success_count': len(self.downloaded_files),
            'failure_count': len(self.failed_urls),
            'success_rate': len(self.downloaded_files) / (len(self.downloaded_files) + len(self.failed_urls)) if (len(self.downloaded_files) + len(self.failed_urls)) > 0 else 0
        }


def main():
    """メイン処理"""
    parser = argparse.ArgumentParser(
        description="JPX上場銘柄一覧ダウンローダー",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
    python scripts/jpx_stock_list_downloader.py
    python scripts/jpx_stock_list_downloader.py --output data/stocks.csv
    python scripts/jpx_stock_list_downloader.py --format csv --validate
    python scripts/jpx_stock_list_downloader.py --output-dir data/jpx --convert-to-csv
        """
    )

    parser.add_argument(
        '--output', '-o',
        type=str,
        help='出力ファイルパス（指定しない場合は自動生成）'
    )

    parser.add_argument(
        '--output-dir',
        type=str,
        default='data',
        help='出力ディレクトリ（デフォルト: data）'
    )

    parser.add_argument(
        '--format',
        choices=['csv', 'excel', 'all'],
        default='all',
        help='ダウンロード形式（デフォルト: all）'
    )

    parser.add_argument(
        '--convert-to-csv',
        action='store_true',
        help='ExcelファイルをCSVに変換'
    )

    parser.add_argument(
        '--validate',
        action='store_true',
        help='ダウンロードファイルの詳細検証を実行'
    )

    parser.add_argument(
        '--timeout',
        type=int,
        default=30,
        help='リクエストタイムアウト秒数（デフォルト: 30）'
    )

    parser.add_argument(
        '--retries',
        type=int,
        default=3,
        help='再試行回数（デフォルト: 3）'
    )

    args = parser.parse_args()

    try:
        # 出力ディレクトリの作成
        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        logger.info("=== JPX銘柄リストダウンローダー開始 ===")

        # ダウンローダーの初期化
        downloader = JPXStockListDownloader(
            timeout=args.timeout,
            retries=args.retries
        )

        # ダウンロード実行
        downloaded_files = downloader.download_all_available_formats(output_dir)

        if not downloaded_files:
            logger.error("ダウンロードに成功したファイルがありません")
            return 1

        # Excel->CSV変換
        if args.convert_to_csv:
            if not PANDAS_AVAILABLE:
                logger.warning("pandasが利用できないため、Excel->CSV変換をスキップします")
            else:
                logger.info("Excel->CSV変換を実行中...")
                for file_path in downloaded_files[:]:  # コピーを作成してイテレート
                    if file_path.suffix.lower() in ['.xls', '.xlsx']:
                        csv_file = downloader.convert_to_csv(file_path)
                        if csv_file:
                            downloaded_files.append(csv_file)

        # サマリー表示
        summary = downloader.get_download_summary()

        logger.info("=== ダウンロード完了 ===")
        logger.info(f"成功: {summary['success_count']}ファイル")
        logger.info(f"失敗: {summary['failure_count']}URL")
        logger.info(f"成功率: {summary['success_rate']:.1%}")

        if downloaded_files:
            logger.info("ダウンロードされたファイル:")
            for file_path in downloaded_files:
                file_size = file_path.stat().st_size / 1024  # KB
                logger.info(f"  - {file_path} ({file_size:.1f} KB)")

        if summary['failed_urls']:
            logger.warning("失敗したURL:")
            for url in summary['failed_urls']:
                logger.warning(f"  - {url}")

        # 特定の出力ファイルが指定された場合
        if args.output and downloaded_files:
            # 最も適切なファイルを選択（CSVファイルを優先）
            csv_files = [f for f in downloaded_files if f.suffix.lower() == '.csv']
            best_file = csv_files[0] if csv_files else downloaded_files[0]

            output_path = Path(args.output)
            if best_file != output_path:
                import shutil
                shutil.copy2(best_file, output_path)
                logger.info(f"出力ファイルをコピー: {best_file} -> {output_path}")

        return 0 if downloaded_files else 1

    except KeyboardInterrupt:
        logger.info("処理が中断されました")
        return 1

    except Exception as e:
        logger.error(f"予期しないエラー: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    exit(main())
