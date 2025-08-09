"""
レポート出力

分析結果・税務データ・パフォーマンスレポートの多形式出力機能
"""

import json
import csv
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Any, Optional
import io

from ...utils.logging_config import get_context_logger

logger = get_context_logger(__name__)


class ReportExporter:
    """
    レポート出力クラス

    各種分析結果を複数形式でエクスポートする機能を提供
    """

    def __init__(self, output_directory: str = "./reports"):
        """
        初期化

        Args:
            output_directory: 出力ディレクトリパス
        """
        self.output_dir = Path(output_directory)
        self.output_dir.mkdir(exist_ok=True)
        logger.info(f"レポート出力機初期化完了 - 出力先: {self.output_dir}")

    def export_portfolio_report(
        self,
        report_data: Dict[str, Any],
        format_type: str = "json",
        filename: Optional[str] = None,
    ) -> str:
        """
        ポートフォリオレポート出力

        Args:
            report_data: レポートデータ
            format_type: 出力形式（json, csv, html）
            filename: ファイル名（指定しない場合は自動生成）

        Returns:
            出力ファイルパス
        """
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"portfolio_report_{timestamp}"

            if format_type.lower() == "json":
                return self._export_to_json(report_data, f"{filename}.json")
            elif format_type.lower() == "csv":
                return self._export_portfolio_to_csv(report_data, f"{filename}.csv")
            elif format_type.lower() == "html":
                return self._export_portfolio_to_html(report_data, f"{filename}.html")
            else:
                raise ValueError(f"未サポートの出力形式: {format_type}")

        except Exception as e:
            logger.error(f"ポートフォリオレポート出力エラー: {e}")
            return ""

    def export_tax_report(
        self,
        tax_data: Dict[str, Any],
        format_type: str = "json",
        filename: Optional[str] = None,
    ) -> str:
        """
        税務レポート出力

        Args:
            tax_data: 税務データ
            format_type: 出力形式
            filename: ファイル名

        Returns:
            出力ファイルパス
        """
        try:
            if filename is None:
                year = tax_data.get("report_info", {}).get("target_year", datetime.now().year)
                filename = f"tax_report_{year}"

            if format_type.lower() == "json":
                return self._export_to_json(tax_data, f"{filename}.json")
            elif format_type.lower() == "csv":
                return self._export_tax_to_csv(tax_data, f"{filename}.csv")
            elif format_type.lower() == "html":
                return self._export_tax_to_html(tax_data, f"{filename}.html")
            else:
                raise ValueError(f"未サポートの出力形式: {format_type}")

        except Exception as e:
            logger.error(f"税務レポート出力エラー: {e}")
            return ""

    def export_trade_transactions(
        self,
        transactions: List[Dict],
        format_type: str = "csv",
        filename: Optional[str] = None,
    ) -> str:
        """
        取引明細出力

        Args:
            transactions: 取引明細リスト
            format_type: 出力形式
            filename: ファイル名

        Returns:
            出力ファイルパス
        """
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"transactions_{timestamp}"

            if format_type.lower() == "csv":
                return self._export_transactions_to_csv(transactions, f"{filename}.csv")
            elif format_type.lower() == "json":
                return self._export_to_json({"transactions": transactions}, f"{filename}.json")
            elif format_type.lower() == "xlsx":
                return self._export_transactions_to_excel(transactions, f"{filename}.xlsx")
            else:
                raise ValueError(f"未サポートの出力形式: {format_type}")

        except Exception as e:
            logger.error(f"取引明細出力エラー: {e}")
            return ""

    def _export_to_json(self, data: Dict[str, Any], filename: str) -> str:
        """JSON形式出力"""
        file_path = self.output_dir / filename

        try:
            # Decimal型をstrに変換するカスタムエンコーダー
            def decimal_encoder(obj):
                if isinstance(obj, Decimal):
                    return str(obj)
                raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=decimal_encoder)

            logger.info(f"JSON出力完了: {file_path}")
            return str(file_path)

        except Exception as e:
            logger.error(f"JSON出力エラー: {e}")
            return ""

    def _export_portfolio_to_csv(self, report_data: Dict[str, Any], filename: str) -> str:
        """ポートフォリオCSV出力"""
        file_path = self.output_dir / filename

        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)

                # ヘッダー情報
                writer.writerow(["ポートフォリオレポート"])
                writer.writerow(["生成日時", datetime.now().strftime("%Y/%m/%d %H:%M")])
                writer.writerow([])

                # パフォーマンスメトリクス
                performance = report_data.get("performance_metrics", {})
                writer.writerow(["パフォーマンス指標"])
                writer.writerow(["項目", "値", "単位"])

                perf_items = [
                    ("総投資額", "total_investment", "円"),
                    ("実現損益", "realized_pnl", "円"),
                    ("未実現損益", "unrealized_pnl", "円"),
                    ("総リターン", "total_return", "円"),
                    ("リターン率", "return_percentage", "%"),
                    ("年率換算リターン", "annualized_return", "%"),
                    ("シャープレシオ", "sharpe_ratio", ""),
                    ("最大ドローダウン", "max_drawdown_percentage", "%"),
                ]

                for label, key, unit in perf_items:
                    value = performance.get(key, "N/A")
                    writer.writerow([label, value, unit])

                writer.writerow([])

                # リスク指標
                risk_metrics = report_data.get("risk_metrics", {})
                writer.writerow(["リスク指標"])
                writer.writerow(["項目", "値", "単位"])

                risk_items = [
                    ("ポートフォリオ価値", "portfolio_value", "円"),
                    ("VaR(95%)", "value_at_risk_95", "円"),
                    ("集中リスク", "concentration_risk_percentage", "%"),
                    ("ポートフォリオベータ", "portfolio_beta", ""),
                    ("年率ボラティリティ", "annualized_volatility", "%"),
                ]

                for label, key, unit in risk_items:
                    value = risk_metrics.get(key, "N/A")
                    writer.writerow([label, value, unit])

            logger.info(f"ポートフォリオCSV出力完了: {file_path}")
            return str(file_path)

        except Exception as e:
            logger.error(f"ポートフォリオCSV出力エラー: {e}")
            return ""

    def _export_tax_to_csv(self, tax_data: Dict[str, Any], filename: str) -> str:
        """税務CSV出力"""
        file_path = self.output_dir / filename

        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)

                # 基本情報
                report_info = tax_data.get("report_info", {})
                writer.writerow(["確定申告用データ"])
                writer.writerow(["対象年度", report_info.get("target_year", "N/A")])
                writer.writerow(["作成日時", report_info.get("generation_date", "N/A")])
                writer.writerow([])

                # 税務サマリー
                tax_summary = tax_data.get("tax_summary", {})
                writer.writerow(["税務サマリー"])
                writer.writerow(["項目", "金額(円)"])

                tax_items = [
                    ("総購入額", "total_buy_amount"),
                    ("総売却額", "total_sell_amount"),
                    ("実現損益", "realized_gains"),
                    ("手数料合計", "total_fees"),
                    ("課税所得", "taxable_income"),
                    ("所得税", "income_tax"),
                    ("住民税", "local_tax"),
                    ("復興特別所得税", "special_reconstruction_tax"),
                    ("税額合計", "total_tax"),
                    ("繰越損失", "loss_carryover"),
                ]

                for label, key in tax_items:
                    value = tax_summary.get(key, "N/A")
                    writer.writerow([label, value])

                writer.writerow([])

                # 銘柄別明細
                stock_details = tax_data.get("stock_details", [])
                if stock_details:
                    writer.writerow(["銘柄別損益明細"])
                    writer.writerow(["銘柄", "購入回数", "売却回数", "購入額", "売却額", "実現損益", "手数料"])

                    for stock in stock_details:
                        writer.writerow([
                            stock["symbol"],
                            stock["buy_transactions"],
                            stock["sell_transactions"],
                            stock["total_buy_amount"],
                            stock["total_sell_amount"],
                            stock["realized_pnl"],
                            stock["total_fees"],
                        ])

            logger.info(f"税務CSV出力完了: {file_path}")
            return str(file_path)

        except Exception as e:
            logger.error(f"税務CSV出力エラー: {e}")
            return ""

    def _export_transactions_to_csv(self, transactions: List[Dict], filename: str) -> str:
        """取引明細CSV出力"""
        file_path = self.output_dir / filename

        try:
            if not transactions:
                logger.warning("取引データが空です")
                return ""

            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                # ヘッダー作成
                headers = list(transactions[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=headers)

                # ヘッダー行出力
                writer.writeheader()

                # データ出力
                for transaction in transactions:
                    # Decimal型を文字列に変換
                    row = {}
                    for key, value in transaction.items():
                        if isinstance(value, Decimal):
                            row[key] = str(value)
                        else:
                            row[key] = value
                    writer.writerow(row)

            logger.info(f"取引明細CSV出力完了: {file_path} ({len(transactions)}件)")
            return str(file_path)

        except Exception as e:
            logger.error(f"取引明細CSV出力エラー: {e}")
            return ""

    def _export_portfolio_to_html(self, report_data: Dict[str, Any], filename: str) -> str:
        """ポートフォリオHTML出力"""
        file_path = self.output_dir / filename

        try:
            html_content = self._generate_portfolio_html(report_data)

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

            logger.info(f"ポートフォリオHTML出力完了: {file_path}")
            return str(file_path)

        except Exception as e:
            logger.error(f"ポートフォリオHTML出力エラー: {e}")
            return ""

    def _generate_portfolio_html(self, report_data: Dict[str, Any]) -> str:
        """ポートフォリオHTML生成"""
        performance = report_data.get("performance_metrics", {})
        risk_metrics = report_data.get("risk_metrics", {})
        efficiency = report_data.get("efficiency_metrics", {})

        html = f"""
        <!DOCTYPE html>
        <html lang="ja">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>ポートフォリオレポート</title>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 20px; }}
                .header {{ background-color: #2c3e50; color: white; padding: 20px; border-radius: 5px; }}
                .section {{ margin: 20px 0; padding: 15px; border: 1px solid #ddd; border-radius: 5px; }}
                .metrics-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; }}
                .metric-card {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; }}
                .positive {{ color: #27ae60; }}
                .negative {{ color: #e74c3c; }}
                table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
                th, td {{ padding: 10px; text-align: right; border-bottom: 1px solid #ddd; }}
                th {{ background-color: #34495e; color: white; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>ポートフォリオレポート</h1>
                <p>生成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}</p>
            </div>

            <div class="section">
                <h2>パフォーマンス指標</h2>
                <div class="metrics-grid">
                    <div class="metric-card">
                        <h3>リターン</h3>
                        <p>総リターン: <span class="{'positive' if float(performance.get('total_return', 0)) >= 0 else 'negative'}">{performance.get('total_return', 'N/A')}円</span></p>
                        <p>リターン率: <span class="{'positive' if float(performance.get('return_percentage', 0)) >= 0 else 'negative'}">{performance.get('return_percentage', 'N/A')}%</span></p>
                        <p>年率リターン: {performance.get('annualized_return', 'N/A')}%</p>
                    </div>
                    <div class="metric-card">
                        <h3>リスク指標</h3>
                        <p>シャープレシオ: {performance.get('sharpe_ratio', 'N/A')}</p>
                        <p>最大ドローダウン: {performance.get('max_drawdown_percentage', 'N/A')}%</p>
                        <p>集中リスク: {risk_metrics.get('concentration_risk_percentage', 'N/A')}%</p>
                    </div>
                    <div class="metric-card">
                        <h3>効率指標</h3>
                        <p>総取引数: {efficiency.get('total_trades', 'N/A')}回</p>
                        <p>平均手数料: {efficiency.get('average_commission', 'N/A')}円</p>
                        <p>手数料率: {efficiency.get('commission_percentage', 'N/A')}%</p>
                    </div>
                </div>
            </div>

            <div class="section">
                <h2>詳細指標</h2>
                <table>
                    <tr><th>指標</th><th>値</th><th>単位</th></tr>
                    <tr><td>総投資額</td><td>{performance.get('total_investment', 'N/A')}</td><td>円</td></tr>
                    <tr><td>実現損益</td><td>{performance.get('realized_pnl', 'N/A')}</td><td>円</td></tr>
                    <tr><td>未実現損益</td><td>{performance.get('unrealized_pnl', 'N/A')}</td><td>円</td></tr>
                    <tr><td>VaR(95%)</td><td>{risk_metrics.get('value_at_risk_95', 'N/A')}</td><td>円</td></tr>
                    <tr><td>ポートフォリオベータ</td><td>{risk_metrics.get('portfolio_beta', 'N/A')}</td><td>-</td></tr>
                </table>
            </div>

            <div class="section">
                <p><small>このレポートは自動生成されました。投資判断は自己責任で行ってください。</small></p>
            </div>
        </body>
        </html>
        """

        return html

    def _export_tax_to_html(self, tax_data: Dict[str, Any], filename: str) -> str:
        """税務HTML出力"""
        file_path = self.output_dir / filename

        try:
            html_content = self._generate_tax_html(tax_data)

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

            logger.info(f"税務HTML出力完了: {file_path}")
            return str(file_path)

        except Exception as e:
            logger.error(f"税務HTML出力エラー: {e}")
            return ""

    def _generate_tax_html(self, tax_data: Dict[str, Any]) -> str:
        """税務HTML生成"""
        report_info = tax_data.get("report_info", {})
        tax_summary = tax_data.get("tax_summary", {})
        stock_details = tax_data.get("stock_details", [])

        # 銘柄別テーブル生成
        stock_table_rows = ""
        for stock in stock_details:
            pnl_class = "positive" if float(stock.get("realized_pnl", 0)) >= 0 else "negative"
            stock_table_rows += f"""
            <tr>
                <td>{stock['symbol']}</td>
                <td>{stock['buy_transactions']}</td>
                <td>{stock['sell_transactions']}</td>
                <td>{stock['total_buy_amount']:,}</td>
                <td>{stock['total_sell_amount']:,}</td>
                <td class="{pnl_class}">{stock['realized_pnl']:,}</td>
                <td>{stock['total_fees']:,}</td>
            </tr>
            """

        html = f"""
        <!DOCTYPE html>
        <html lang="ja">
        <head>
            <meta charset="UTF-8">
            <title>確定申告用データ - {report_info.get('target_year', '')}年</title>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 20px; }}
                .header {{ background-color: #34495e; color: white; padding: 20px; border-radius: 5px; }}
                .section {{ margin: 20px 0; padding: 15px; border: 1px solid #ddd; border-radius: 5px; }}
                .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 15px; }}
                .summary-item {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; }}
                .positive {{ color: #27ae60; font-weight: bold; }}
                .negative {{ color: #e74c3c; font-weight: bold; }}
                table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
                th, td {{ padding: 10px; text-align: right; border-bottom: 1px solid #ddd; }}
                th {{ background-color: #2c3e50; color: white; }}
                .important {{ background-color: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>確定申告用データ</h1>
                <p>対象年度: {report_info.get('target_year', 'N/A')}年</p>
                <p>作成日時: {report_info.get('generation_date', 'N/A')}</p>
            </div>

            <div class="section important">
                <h2>重要な税務情報</h2>
                <div class="summary-grid">
                    <div class="summary-item">
                        <h3>課税所得</h3>
                        <p style="font-size: 1.5em;"><span class="{'positive' if float(tax_summary.get('taxable_income', 0)) >= 0 else 'negative'}">{tax_summary.get('taxable_income', 'N/A'):,}円</span></p>
                    </div>
                    <div class="summary-item">
                        <h3>納税額合計</h3>
                        <p style="font-size: 1.5em;">{tax_summary.get('total_tax', 'N/A'):,}円</p>
                    </div>
                    <div class="summary-item">
                        <h3>繰越損失</h3>
                        <p style="font-size: 1.5em;">{tax_summary.get('loss_carryover', 'N/A'):,}円</p>
                    </div>
                </div>
            </div>

            <div class="section">
                <h2>税務サマリー詳細</h2>
                <table>
                    <tr><th>項目</th><th>金額</th></tr>
                    <tr><td>総購入額</td><td>{tax_summary.get('total_buy_amount', 'N/A'):,}円</td></tr>
                    <tr><td>総売却額</td><td>{tax_summary.get('total_sell_amount', 'N/A'):,}円</td></tr>
                    <tr><td>実現損益</td><td class="{'positive' if float(tax_summary.get('realized_gains', 0)) >= 0 else 'negative'}">{tax_summary.get('realized_gains', 'N/A'):,}円</td></tr>
                    <tr><td>手数料合計</td><td>{tax_summary.get('total_fees', 'N/A'):,}円</td></tr>
                    <tr><td>所得税</td><td>{tax_summary.get('income_tax', 'N/A'):,}円</td></tr>
                    <tr><td>住民税</td><td>{tax_summary.get('local_tax', 'N/A'):,}円</td></tr>
                    <tr><td>復興特別所得税</td><td>{tax_summary.get('special_reconstruction_tax', 'N/A'):,}円</td></tr>
                </table>
            </div>

            <div class="section">
                <h2>銘柄別損益明細</h2>
                <table>
                    <tr>
                        <th>銘柄</th>
                        <th>購入回数</th>
                        <th>売却回数</th>
                        <th>購入額(円)</th>
                        <th>売却額(円)</th>
                        <th>実現損益(円)</th>
                        <th>手数料(円)</th>
                    </tr>
                    {stock_table_rows}
                </table>
            </div>

            <div class="section">
                <p><small>このデータは確定申告の参考用です。正式な申告前に税理士等にご相談ください。</small></p>
            </div>
        </body>
        </html>
        """

        return html

    def _export_transactions_to_excel(self, transactions: List[Dict], filename: str) -> str:
        """Excel形式出力（要openpyxl）"""
        try:
            # openpyxlが利用可能かチェック
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            logger.error("Excel出力にはopenpyxlが必要です")
            return ""

        file_path = self.output_dir / filename

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "取引明細"

            # ヘッダースタイル設定
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # ヘッダー行
            if transactions:
                headers = list(transactions[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border

                # データ行
                for row, transaction in enumerate(transactions, 2):
                    for col, (key, value) in enumerate(transaction.items(), 1):
                        cell = ws.cell(row=row, column=col, value=str(value))
                        cell.border = border

                        # 数値の場合は右寄せ
                        if isinstance(value, (int, float, Decimal)):
                            cell.alignment = Alignment(horizontal='right')

            # 列幅自動調整
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            logger.info(f"Excel出力完了: {file_path}")
            return str(file_path)

        except Exception as e:
            logger.error(f"Excel出力エラー: {e}")
            return ""

    def export_multiple_formats(
        self,
        data: Dict[str, Any],
        data_type: str,
        base_filename: Optional[str] = None,
        formats: List[str] = None,
    ) -> List[str]:
        """
        複数形式一括出力

        Args:
            data: 出力データ
            data_type: データ種類（portfolio, tax, transactions）
            base_filename: ベースファイル名
            formats: 出力形式リスト

        Returns:
            出力ファイルパスリスト
        """
        if formats is None:
            formats = ["json", "csv", "html"]

        output_files = []

        for format_type in formats:
            try:
                if data_type == "portfolio":
                    file_path = self.export_portfolio_report(data, format_type, base_filename)
                elif data_type == "tax":
                    file_path = self.export_tax_report(data, format_type, base_filename)
                elif data_type == "transactions":
                    file_path = self.export_trade_transactions(data, format_type, base_filename)
                else:
                    logger.error(f"未知のデータタイプ: {data_type}")
                    continue

                if file_path:
                    output_files.append(file_path)

            except Exception as e:
                logger.error(f"{format_type}形式出力エラー: {e}")

        logger.info(f"複数形式出力完了: {len(output_files)}ファイル")
        return output_files

    def get_output_directory(self) -> str:
        """出力ディレクトリパス取得"""
        return str(self.output_dir)

    def clean_old_reports(self, days_old: int = 30) -> int:
        """
        古いレポートファイル削除

        Args:
            days_old: 削除対象日数

        Returns:
            削除ファイル数
        """
        try:
            cutoff_date = datetime.now() - timedelta(days=days_old)
            deleted_count = 0

            for file_path in self.output_dir.glob("*"):
                if file_path.is_file():
                    file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if file_mtime < cutoff_date:
                        file_path.unlink()
                        deleted_count += 1
                        logger.debug(f"古いレポート削除: {file_path}")

            logger.info(f"古いレポートクリーンアップ完了: {deleted_count}ファイル削除")
            return deleted_count

        except Exception as e:
            logger.error(f"レポートクリーンアップエラー: {e}")
            return 0
